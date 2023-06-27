import glob
import sqlite3
import openpyxl
import warnings

warnings.simplefilter("ignore")


dict_for_operator = \
    {
        'Общество с ограниченной ответственностью «Скартел»': 'Скартел',
        'Общество с ограниченной ответственностью \"Скартел\"': 'Скартел',

        'Общество с ограниченной ответственностью \"Т2 Мобайл\"': 'Т2 Мобайл',
        'Общество с ограниченной ответственностью «Т2 Мобайл»': 'Т2 Мобайл',

        'Публичное акционерное общество «Мобильные ТелеСистемы»': 'МТС',
        'Публичное акционерное общество \"Мобильные ТелеСистемы\"': 'МТС',

        'Публичное акционерное общество \"МегаФон\"': 'МегаФон',
        'Публичное акционерное общество «МегаФон»': 'МегаФон',

        'Публичное акционерное общество \"Ростелеком\"': 'Ростелеком',
        'Публичное акционерное общество «Ростелеком»': 'Ростелеком',
        'Публичное акционерное общество междугородной и международной электрической связи \"Ростелеком\"': 'Ростелеком',

        'Публичное акционерное общество «Вымпел-Коммуникации»': 'ВымпелКом',
        'Публичное акционерное общество \"Вымпел-Коммуникации\"': 'ВымпелКом'
    }

dict_ETC = {
    '18.1.1.3.': 'GSM',
    '18.1.1.8.': 'GSM',
    '18.1.1.5.': 'UMTS',
    '18.1.1.6.': 'UMTS',
    '18.7.1.': 'LTE',
    '18.7.4.': 'LTE',
    '18.7.5.': '5G NR',
    '19.2.': 'РРС'
}

def measure_time(func):
    def wrapper(*args, **kwargs):
        start_time = datetime.datetime.now()
        result = func(*args, **kwargs)
        end_time = datetime.datetime.now()
        elapsed_time = end_time - start_time
        print(f"Execution time: {elapsed_time}")
        return result

    return wrapper

@measure_time
def export_to_sqlite(file_open):
    base_name = 'from_eirs.sqlite3'

    # метод sqlite3.connect автоматически создаст базу, если ее нет
    connect = sqlite3.connect('data_bases\\' + base_name)

    cursor = connect.cursor()

    cursor.execute('CREATE TABLE IF NOT EXISTS cellular (РЭС, Адрес, ТИП_РЭС, Состояние, Владелец, Широта, Долгота, Частоты, Дополнительные_параметры, Классы_излучения, Серия_Номер_РЗ_СоР)')

    file_to_read = openpyxl.load_workbook(file_open, data_only=True)
    sheet = file_to_read['SQL Results']

    for row in range(2, sheet.max_row + 1):
        data = []

        for col in range(1, sheet.max_column + 1):
            value = sheet.cell(row, col).value
            data.append(value)
        if data[5] == 'Закрытое':
            continue
        cursor.execute("INSERT INTO cellular VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", (data[1], data[2], dict_ETC[data[3]], data[5], dict_for_operator[data[6]], data[7], data[8], data[10], data[11], data[17], f'{data[18]} {data[19]}'))

    connect.commit()
    connect.close()

file_xlxl_1 = glob.glob('source_folder\*.xlsx')

export_to_sqlite(file_xlxl_1[0])